from openpyxl import Workbook, load_workbook
wb = load_workbook('excel\sample3.xlsx')

# grab the active worksheet
ws = wb.active
#acha a ultima linha
for cell in ws['c']:
    if cell.value != '':
        linha = cell.row
        linha += 1

ws[f'c{linha}'] = 'victor'
# Save the file
wb.save("excel\sample3.xlsx")
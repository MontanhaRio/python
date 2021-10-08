from openpyxl import Workbook, load_workbook
wb = load_workbook('samplepy.xlsx')

# grab the active worksheet
ws = wb.active

for cell in ws['c']:
    if cell.value == 'carro':
        linha = cell.row
        ws[f'd{linha}'] = 20
        ws[f'e{linha}'] = 33
# Save the file
wb.save("samplepy.xlsx")
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Crie a pasta de trabalho e a planilha com as quais trabalharemos
wb = Workbook()
ws = wb.active

# Crie um objeto de validação de dados com validação de lista
dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)

# Opcionalmente, defina uma mensagem de erro personalizada
dv.error ='Sua entrada não está na lista'
dv.errorTitle = 'Entrada inválida'

# Opcionalmente, defina uma mensagem de prompt personalizada
dv.prompt = 'Selecione na lista'
dv.promptTitle = 'Lista de seleção'

# Adicione o objeto de validação de dados à planilha
ws.add_data_validation(dv)

# Crie algumas células e adicione-as ao objeto de validação de dados
c1 = ws["A1"]
c1.value = "Dog"
dv.add(c1)
c2 = ws["A2"]
c2.value = "Um valor inválido"
dv.add(c2)

# Ou aplique a validação a um intervalo de células
dv.add('B1:B1048576')# Este é o mesmo que para toda a coluna

# Verifique se uma célula está no validador
"B4" in dv
True
wb.save('vd.xlsx')
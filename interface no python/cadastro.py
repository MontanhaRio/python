import PySimpleGUI as sg
from PySimpleGUI.PySimpleGUI import popup 
from openpyxl import Workbook, load_workbook
wb = Workbook()
wb = load_workbook('python.xlsx')           
sh = wb.worksheets[0]
class TelaPython:
    def __init__(self):
        #Layout
        layout = [
            [sg.Text('Nome',size=(5,0)),sg.Input(size=(15,0),key='nome')],
            [sg.Text('Idade',size=(5,0)),sg.Input(size=(15,0),key='idade')],
            [sg.Text('Quais provedores de e-mail sao aceitos?')],
            [sg.Checkbox('Gmail',key='gmail'),sg.Checkbox('Outlook',key='outlook'),sg.Checkbox('Yahoo',key='yahoo')],
            [sg.Text('Aceita cartao')],
            [sg.Radio('Sim','cartoes',key='aceitaCartao'),sg.Radio('Nao','cartoes',key='NaoAceitaCartao')],
            [sg.Combo(sh['A'], size=(20,0),key='dados')],
            [sg.Button('Enviar Dados')],
            [sg.Output(size=(30,20))],
            
        ]
        # Janela
        self.janela = sg.Window("Dados do Usuario",layout=layout)
        # Extrair os dados da tela
        
    def Iniciar(self):
        while True:
            self.button, self.values = self.janela.Read()
            nome = self.values['nome']
            idade = self.values['idade']
            aceita_gmail = self.values['gmail']
            aceita_outlook =self.values['outlook']
            aceita_yahoo = self.values['yahoo']
            aceita_cartao = self.values['aceitaCartao']
            nao_aceita_cartao = self.values['NaoAceitaCartao']
            lista = self.values['dodas']
            print(f'Nome: {nome}')
            print(f'Idade: {idade}')
            print(f'Aceita gmail: {aceita_gmail}')
            print(f'Aceita outlook: {aceita_outlook}')
            print(f'Aceita yahoo: {aceita_yahoo}')
            print(f'Aceita Cartao: {aceita_cartao}')
            print(f'Nao aceita Cartao: {nao_aceita_cartao}')
            print(f'Dentro do excel: {lista}')
    

            
            
            sh.title = 'cadastro'
            sh['A1'] = "Nomes"
            sh['B1'] = "Idades"
            for cell in sh['A']:
                if cell.value != '':
                    linha = cell.row
                    linha += 1
            sh[f'A{linha}'] = self.values['nome']
            sh[f'B{linha}'] = self.values['idade']
            

            
                    
                
            wb.save('python.xlsx')
            
cadastro = TelaPython()
cadastro.Iniciar()
